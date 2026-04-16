import asyncio
import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

SEVERITY_TH = {
    "normal":     "ปกติ",
    "none":       "ไม่มีอาการ",
    "mild":       "น้อย",
    "moderate":   "ปานกลาง",
    "severe":     "มาก",
    "very_severe":"รุนแรงมาก",
    "clinical":   "ต้องดูแล",
}


async def get_report_data(db, current_user, filters: dict) -> list[dict]:
    from sqlalchemy import text
    from app.deps import check_report_scope

    scope = await check_report_scope(
        current_user,
        filters.get("school_id"),
        filters.get("district_id"),
        filters.get("affiliation_id"),
    )

    query = """
    SELECT
        s.first_name, s.last_name, s.grade, s.gender,
        sc.name AS school_name,
        a.assessment_type, a.score, a.severity_level,
        a.suicide_risk, a.created_at
    FROM assessments a
    JOIN students s ON a.student_id = s.id
    LEFT JOIN schools sc ON s.school_id = sc.id
    LEFT JOIN districts d ON sc.district_id = d.id
    WHERE 1=1
    """
    params: dict = {}

    if scope.school_id:
        query += " AND s.school_id = :school_id"
        params["school_id"] = scope.school_id
    if scope.district_id:
        query += " AND sc.district_id = :district_id"
        params["district_id"] = scope.district_id
    if scope.affiliation_id:
        query += " AND d.affiliation_id = :affiliation_id"
        params["affiliation_id"] = scope.affiliation_id

    if filters.get("assessment_type"):
        query += " AND a.assessment_type = :assessment_type"
        params["assessment_type"] = filters["assessment_type"]
    if filters.get("grade"):
        query += " AND s.grade = :grade"
        params["grade"] = filters["grade"]
    if filters.get("gender"):
        query += " AND s.gender = :gender"
        params["gender"] = filters["gender"]
    if filters.get("date_from"):
        query += " AND a.created_at >= :date_from"
        params["date_from"] = datetime.fromisoformat(str(filters["date_from"]))
    if filters.get("date_to"):
        query += " AND a.created_at <= :date_to"
        params["date_to"] = datetime.fromisoformat(str(filters["date_to"])).replace(
            hour=23, minute=59, second=59
        )

    query += " ORDER BY a.created_at DESC LIMIT 5000"

    result = await db.execute(text(query), params)
    return [dict(row._mapping) for row in result.fetchall()]


# ─── Excel ────────────────────────────────────────────────────────────────────

async def generate_excel_report(db, current_user, filters: dict) -> bytes:
    data = await get_report_data(db, current_user, filters)

    wb = Workbook()
    ws = wb.active
    ws.title = "รายงาน LEMCS"

    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center")

    headers = [
        "ลำดับ", "ชื่อ", "นามสกุล", "โรงเรียน", "ระดับชั้น", "เพศ",
        "ประเภทแบบประเมิน", "คะแนน", "ระดับความเสี่ยง",
        "ความเสี่ยงการฆ่าตัวตาย", "วันที่ประเมิน",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    for i, row in enumerate(data, 2):
        ws.cell(row=i, column=1,  value=i - 1)
        ws.cell(row=i, column=2,  value=row.get("first_name", ""))
        ws.cell(row=i, column=3,  value=row.get("last_name", ""))
        ws.cell(row=i, column=4,  value=row.get("school_name") or "ไม่มีข้อมูล")
        ws.cell(row=i, column=5,  value=row.get("grade", ""))
        ws.cell(row=i, column=6,  value=row.get("gender", ""))
        ws.cell(row=i, column=7,  value=row.get("assessment_type", ""))
        ws.cell(row=i, column=8,  value=row.get("score"))
        ws.cell(row=i, column=9,  value=SEVERITY_TH.get(row.get("severity_level", ""), row.get("severity_level", "")))
        ws.cell(row=i, column=10, value="ใช่" if row.get("suicide_risk") else "ไม่มี")
        created = row.get("created_at")
        ws.cell(row=i, column=11, value=created.strftime("%d/%m/%Y %H:%M") if created else "")

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 45)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


# ─── PDF ──────────────────────────────────────────────────────────────────────

def _build_pdf_html(data: list[dict]) -> str:
    rows_html = ""
    for i, row in enumerate(data, 1):
        risk = row.get("suicide_risk")
        td_risk = f'<td style="color:#dc2626;font-weight:700">{"ใช่" if risk else "ไม่มี"}</td>'
        created = row.get("created_at")
        date_str = created.strftime("%d/%m/%Y") if created else ""
        rows_html += f"""<tr class="{'even' if i % 2 == 0 else ''}">
            <td style="text-align:center">{i}</td>
            <td>{row.get('first_name','')} {row.get('last_name','')}</td>
            <td>{row.get('school_name') or ''}</td>
            <td style="text-align:center">{row.get('grade','')}</td>
            <td style="text-align:center">{row.get('gender','')}</td>
            <td style="text-align:center">{row.get('assessment_type','')}</td>
            <td style="text-align:center">{row.get('score','')}</td>
            <td>{SEVERITY_TH.get(row.get('severity_level',''), row.get('severity_level',''))}</td>
            {td_risk}
            <td style="text-align:center">{date_str}</td>
        </tr>"""

    generated = datetime.now().strftime("%d/%m/%Y %H:%M")
    return f"""<!DOCTYPE html>
<html lang="th">
<head>
<meta charset="UTF-8">
<style>
  body {{
    font-family: 'Noto Sans Thai', 'Noto Sans', sans-serif;
    font-size: 11px;
    margin: 1.5cm 2cm;
    color: #111;
  }}
  h1 {{ font-size: 16px; text-align: center; color: #1d4ed8; margin-bottom: 4px; }}
  .sub {{ text-align: center; color: #6b7280; font-size: 10px; margin-bottom: 16px; }}
  table {{ width: 100%; border-collapse: collapse; margin-top: 8px; }}
  th {{
    background: #1d4ed8; color: #fff; padding: 6px 8px;
    text-align: left; font-size: 10px;
  }}
  td {{ padding: 5px 8px; border-bottom: 1px solid #e5e7eb; font-size: 10px; }}
  tr.even td {{ background: #f8fafc; }}
  .footer {{
    text-align: center; color: #9ca3af; font-size: 9px;
    margin-top: 20px; border-top: 1px solid #e5e7eb; padding-top: 8px;
  }}
</style>
</head>
<body>
  <h1>LEMCS — รายงานผลการประเมินสุขภาพจิต</h1>
  <p class="sub">ระบบคัดกรองสุขภาพจิตนักเรียน จังหวัดเลย</p>
  <table>
    <thead>
      <tr>
        <th>#</th><th>ชื่อ-นามสกุล</th><th>โรงเรียน</th>
        <th>ชั้น</th><th>เพศ</th><th>แบบประเมิน</th>
        <th>คะแนน</th><th>ระดับ</th><th>เสี่ยงฆ่าตัวตาย</th><th>วันที่</th>
      </tr>
    </thead>
    <tbody>{rows_html}</tbody>
  </table>
  <p class="footer">สร้างเมื่อ {generated} | LEMCS Loei Educational MindCare System</p>
</body>
</html>"""


async def generate_pdf_report(db, current_user, filters: dict) -> bytes:
    data = await get_report_data(db, current_user, filters)
    html = _build_pdf_html(data)

    # weasyprint เป็น sync/blocking — รันใน thread pool เพื่อไม่ block event loop
    def _render():
        from weasyprint import HTML, CSS
        return HTML(string=html).write_pdf()

    return await asyncio.to_thread(_render)
