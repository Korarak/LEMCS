"""
Smart Import Service: Excel → Bulk DB Insert (ไม่ต้องใช้ Template)
รองรับการ auto-detect header row และ fuzzy column mapping
รองรับ: students-loeitech.xlsx (ระบบอาชีวะ) และ students-เชียงกลม.xlsx (SGS/OBEC)
"""
import io
import csv
import re
import traceback
from difflib import SequenceMatcher
from typing import Any
from datetime import date, datetime
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from app.models.db_models import Student, School, District, User
from app.services.encryption import encrypt_pii, hash_pii


# ─── Column Alias Map ────────────────────────────────────────────────────────
# ชื่อ field ในระบบ → รายชื่อ alias ที่อาจพบในไฟล์จากโรงเรียนต่างๆ
COLUMN_ALIASES: dict[str, list[str]] = {
    "national_id": [
        "เลขประจำตัวประชาชน",
        "เลขบัตรปชช",
        "เลขบัตรประชาชน",
        "รหัสบัตรประชาชน",
        "national_id",
        "เลขประจำตัวประชาชน13หลัก",
    ],
    "student_code": [
        "รหัสประจำตัว",
        "รหัสนักเรียน",
        "student_code",
        "เลขประจำตัวนักเรียน",
        "รหัสนร.",
        "เลขประจำตัว",
    ],
    "prefix": [
        "คำนำหน้าชื่อ", "คำนำหน้า", "prefix", "นำหน้า",
        "ชื่อ - นามสกุล",  # loeitech ใช้ column นี้เป็น prefix (ชื่อ-นามสกุลอยู่ col ถัดไป)
    ],
    "first_name": [
        "ชื่อ", "first_name", "ชื่อจริง", "ชื่อนักเรียน",
    ],
    "last_name": [
        "นามสกุล", "last_name", "สกุล",
    ],
    "gender": [
        "เพศ", "gender", "sex",
    ],
    "birthdate": [
        "วันเกิด", "ว.ด.ป. เกิด", "birthdate", "วันเดือนปีเกิด",
        "วันเกิด(พ.ศ.)", "วันที่เกิด",
    ],
    "grade": [
        "ชั้น", "ระดับชั้น", "grade", "กลุ่มเรียน", "ชั้นปี",
        "ระดับ", "ชั้น/ปี",
    ],
    "classroom": [
        "ห้อง", "classroom", "รหัสกลุ่มเรียน", "ห้องเรียน", "กลุ่ม",
    ],
    "school_name": [
        "ชื่อโรงเรียน", "school_name", "โรงเรียน", "สถาบัน",
    ],
    "school_code": [
        "รหัสโรงเรียน", "school_code", "รหัสสถานศึกษา",
    ],
    "status": [
        "สถานะนักเรียน", "สถานะ", "status",
    ],
}

# ─── Thai Month Map & Date Parsing ───────────────────────────────────────────
THAI_MONTHS = {
    "ม.ค.": 1,  "ก.พ.": 2,  "มี.ค.": 3, "เม.ย.": 4,
    "พ.ค.": 5,  "มิ.ย.": 6, "ก.ค.": 7,  "ส.ค.": 8,
    "ก.ย.": 9,  "ต.ค.": 10, "พ.ย.": 11, "ธ.ค.": 12,
    "มกราคม": 1, "กุมภาพันธ์": 2, "มีนาคม": 3, "เมษายน": 4,
    "พฤษภาคม": 5, "มิถุนายน": 6, "กรกฎาคม": 7, "สิงหาคม": 8,
    "กันยายน": 9, "ตุลาคม": 10, "พฤศจิกายน": 11, "ธันวาคม": 12,
}

MALE_PREFIXES = {"นาย", "เด็กชาย", "ด.ช.", "mr.", "mr"}
FEMALE_PREFIXES = {"นาง", "นางสาว", "น.ส.", "เด็กหญิง", "ด.ญ.", "mrs.", "ms.", "miss"}

# แปลง alias → คำนำหน้ามาตรฐาน
_TITLE_MAP: dict[str, str] = {
    "นาย": "นาย", "mr": "นาย", "mr.": "นาย",
    "เด็กชาย": "เด็กชาย", "ด.ช.": "เด็กชาย", "ดช.": "เด็กชาย", "ดช": "เด็กชาย",
    "นาง": "นาง", "mrs": "นาง", "mrs.": "นาง",
    "นางสาว": "นางสาว", "น.ส.": "นางสาว", "นส.": "นางสาว", "นส": "นางสาว",
    "miss": "นางสาว", "ms": "นางสาว", "ms.": "นางสาว",
    "เด็กหญิง": "เด็กหญิง", "ด.ญ.": "เด็กหญิง", "ดญ.": "เด็กหญิง", "ดญ": "เด็กหญิง",
}
_MALE_TITLES   = {"นาย", "เด็กชาย"}
_FEMALE_TITLES = {"นางสาว", "นาง", "เด็กหญิง"}


def normalize_title(raw: str) -> str:
    """แปลง alias คำนำหน้า → ค่ามาตรฐาน (เด็กชาย / เด็กหญิง / นาย / นางสาว / นาง)"""
    t = str(raw).strip()
    return _TITLE_MAP.get(t.lower(), _TITLE_MAP.get(t, t))


def parse_thai_date(s: str) -> date | None:
    """
    แปลง date string หลายรูปแบบ → date object
    - "06 ต.ค. 51"   → 2008-10-06  (loeitech: พ.ศ. 2 หลัก)
    - "06/09/2563"    → 2020-09-06  (เชียงกลม: dd/mm/พ.ศ. 4 หลัก)
    - "2010-01-15"    → 2010-01-15  (ISO format ค.ศ.)
    """
    if not s or str(s).strip() in ("-", "", "None", "null"):
        return None
    s = str(s).strip()

    # ISO format ค.ศ.: "2010-01-15"
    if re.match(r"^\d{4}-\d{2}-\d{2}$", s):
        try:
            return datetime.strptime(s, "%Y-%m-%d").date()
        except Exception:
            return None

    # dd/mm/YYYY (พ.ศ.)
    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})$", s)
    if m:
        d_, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if y > 2400:  # พ.ศ. → ค.ศ.
            y -= 543
        try:
            return date(y, mo, d_)
        except Exception:
            return None

    # "06 ต.ค. 51" หรือ "06 ตุลาคม 2551"
    for th_month, month_num in THAI_MONTHS.items():
        pattern = rf"^(\d{{1,2}})\s*{re.escape(th_month)}\s*(\d{{2,4}})$"
        m2 = re.match(pattern, s)
        if m2:
            d_ = int(m2.group(1))
            y = int(m2.group(2))
            if y < 100:  # พ.ศ. 2 หลัก เช่น 51 → 2551
                y += 2500
            if y > 2400:  # พ.ศ. 4 หลัก
                y -= 543
            try:
                return date(y, month_num, d_)
            except Exception:
                return None

    # Excel numeric date (days since 1900-01-01)
    try:
        from openpyxl.utils.datetime import from_excel
        return from_excel(float(s))
    except Exception:
        pass

    return None


def detect_gender_from_prefix(prefix: str) -> str:
    """อนุมาน gender จากคำนำหน้า สำหรับไฟล์ที่ไม่มี column เพศ"""
    p = str(prefix).strip().lower()
    if p in MALE_PREFIXES:
        return "ชาย"
    if p in FEMALE_PREFIXES:
        return "หญิง"
    return "ไม่ระบุ"


def detect_gender_from_title(title: str) -> str:
    """อนุมาน gender จากคำนำหน้ามาตรฐาน"""
    if title in _MALE_TITLES:
        return "ชาย"
    if title in _FEMALE_TITLES:
        return "หญิง"
    return "ไม่ระบุ"


def normalize_national_id(raw: str) -> tuple[str, str | None]:
    """
    Normalize และ validate national_id
    รับได้ 2 รูปแบบ:
      - เลขบัตรประชาชนไทย: ตัวเลข 13 หลัก (อาจมี dash/space)
      - G-Code (DMC): G ตามด้วยตัวเลข 12 หลัก เช่น G123456789012
        → ออกโดยกระทรวงศึกษาฯ สำหรับนักเรียนไร้รัฐ/ไร้สัญชาติ
    Returns:
      (normalized_id, None)      — valid
      ("", error_message)        — invalid, ไม่ควรบันทึก
    """
    s = raw.strip()
    if not s:
        return "", None

    # G-Code: G/g + 12 digits
    if re.match(r"^[Gg]\d{12}$", s):
        return s.upper(), None  # normalize เป็น uppercase G

    # ลบ dash/space/dot แล้วตรวจว่าเป็น 13 ตัวเลข
    digits = re.sub(r"[\s\-\.]", "", s)
    if re.match(r"^\d{13}$", digits):
        return digits, None

    # ไม่ผ่าน — ระบุสาเหตุ
    if re.match(r"^[Gg]", s):
        actual = len(re.sub(r"[^0-9]", "", s))
        return "", f"G-Code ไม่ถูกรูปแบบ (G + {actual} หลัก, ต้องการ G + 12 หลัก)"
    total = len(re.sub(r"\D", "", s))
    return "", f"เลขบัตรปชช. มี {total} หลัก (ต้องการ 13 หลัก)"


# คำนำหน้าชื่อสถานศึกษาที่ถือว่า "มีอยู่แล้ว" — ไม่ต้องเติมเพิ่ม
_SCHOOL_KNOWN_PREFIXES = (
    "โรงเรียน",
    "วิทยาลัย",       # วิทยาลัยเทคนิค / วิทยาลัยอาชีวศึกษา / วิทยาลัยการอาชีพ
    "มหาวิทยาลัย",
    "สำนักงาน",
    "ศูนย์",
    "สถาบัน",
    "วิทยาเขต",
    "สกร.",
)

# school_type → คำนำหน้าที่ควรเติม
_PREFIX_BY_TYPE: dict[str, str] = {
    "อาชีวศึกษา": "วิทยาลัย",
    "ประถมศึกษา": "โรงเรียน",
    "มัธยมศึกษา": "โรงเรียน",
    "เอกชน":      "โรงเรียน",
    "สกร.":        "สำนักงานส่งเสริมการเรียนรู้",
}


def normalize_school_name(name: str, school_type: str | None = None) -> str:
    """
    เติมคำนำหน้าชื่อสถานศึกษาอัตโนมัติ ถ้าชื่อยังไม่มีคำนำหน้า
    เช่น "เทคนิคเลย" + อาชีวศึกษา → "วิทยาลัยเทคนิคเลย"
         "เลยพิทยาคม" + มัธยมศึกษา → "โรงเรียนเลยพิทยาคม"
         "วิทยาลัยเทคนิคเลย" (มีอยู่แล้ว) → "วิทยาลัยเทคนิคเลย"
    """
    name = name.strip()
    if not name:
        return name
    # ถ้ามีคำนำหน้าแล้วไม่ต้องเติม
    if any(name.startswith(p) for p in _SCHOOL_KNOWN_PREFIXES):
        return name
    prefix = _PREFIX_BY_TYPE.get(school_type or "", "โรงเรียน")
    return f"{prefix}{name}"


def normalize_gender(g: str) -> str:
    """แปลง gender ให้เป็นมาตรฐาน"""
    g = str(g).strip()
    if g in ("ช", "ชาย", "ช.", "M", "m", "male"):
        return "ชาย"
    if g in ("ญ", "หญิง", "ญ.", "F", "f", "female"):
        return "หญิง"
    return "ไม่ระบุ"


# ─── Smart Header Detection ───────────────────────────────────────────────────

def _similarity(a: str, b: str) -> float:
    return SequenceMatcher(None, a.lower().strip(), b.lower().strip()).ratio()


def _score_row_as_header(row: list[Any]) -> float:
    """ให้คะแนนว่าแถวนี้น่าจะเป็น header row มากแค่ไหน"""
    all_aliases = [alias for aliases in COLUMN_ALIASES.values() for alias in aliases]
    score = 0
    cells = [str(c).strip() for c in row if c is not None]
    for cell in cells:
        best = max((_similarity(cell, alias) for alias in all_aliases), default=0)
        if best > 0.7:
            score += 1
    # bonus: ถ้า cell ไม่ใช่ตัวเลขล้วน
    non_numeric = sum(1 for c in cells if not re.match(r"^\d+\.?\d*$", c))
    return score + non_numeric * 0.1


def auto_detect_header_row(rows: list[list[Any]], max_scan: int = 10) -> int:
    """
    หาแถว header โดยอัตโนมัติ
    Returns: index ของ header row (0-indexed)
    Default: แถว 0 ถ้าไม่เจอ
    """
    best_score = 0
    best_idx = 0
    for i, row in enumerate(rows[:max_scan]):
        score = _score_row_as_header(row)
        if score > best_score:
            best_score = score
            best_idx = i
    return best_idx


def map_columns(headers: list[str]) -> dict[str, int | None]:
    """
    จับคู่ header column → field ในระบบ
    Returns: {field_name: col_index หรือ None ถ้าไม่เจอ}
    ใช้ priority: field ที่อยู่ใน COLUMN_ALIASES ก่อน = มี priority สูงกว่า
    """
    mapping: dict[str, int | None] = {field: None for field in COLUMN_ALIASES}
    # best scores per column (เพื่อ one-to-one mapping)
    col_best: dict[int, tuple[float, str]] = {}  # col_idx → (score, field)

    for col_idx, header in enumerate(headers):
        h = str(header).strip()
        if not h:
            continue
        best_field = None
        best_score = 0.0
        for field, aliases in COLUMN_ALIASES.items():
            for alias in aliases:
                score = _similarity(h, alias)
                if score > best_score and score > 0.7:
                    best_score = score
                    best_field = field
        if best_field:
            prev = col_best.get(col_idx)
            if prev is None or best_score > prev[0]:
                col_best[col_idx] = (best_score, best_field)

    # Assign: each column goes to best-scoring field (first-come wins if tie)
    for col_idx, (score, field) in sorted(col_best.items(), key=lambda x: -x[1][0]):
        if mapping[field] is None:
            mapping[field] = col_idx

    # ── Heuristic: ชื่อ-นามสกุลอยู่ถัด prefix (loeitech pattern) ──────────────
    # ถ้า prefix detect ได้ แต่ first_name / last_name ไม่ detect
    # → สมมุติว่า first_name = prefix+1, last_name = prefix+2
    if mapping["prefix"] is not None:
        p_idx = mapping["prefix"]
        if mapping["first_name"] is None and (p_idx + 1) < len(headers):
            # col ถัดไปต้องไม่ถูก assign แล้ว
            occupied = set(v for v in mapping.values() if v is not None)
            if (p_idx + 1) not in occupied:
                mapping["first_name"] = p_idx + 1
        if mapping["last_name"] is None and (p_idx + 2) < len(headers):
            occupied = set(v for v in mapping.values() if v is not None)
            if (p_idx + 2) not in occupied:
                mapping["last_name"] = p_idx + 2

    return mapping


# ─── Core: Parse Excel/CSV ────────────────────────────────────────────────────

def parse_csv(content: bytes) -> list[dict]:
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    return [row for row in reader]


def _read_excel_raw(content: bytes) -> tuple[str | None, list[list[Any]]]:
    """อ่าน Excel → (sheet_name, all_rows as list of list)"""
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl not installed. Run: pip install openpyxl")
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    rows = [list(row) for row in ws.iter_rows(values_only=True)]
    return ws.title, rows


def _detect_school_from_title_rows(all_rows: list[list[Any]], header_idx: int) -> str:
    """
    สแกนแถวก่อน header row หาชื่อสถานศึกษา
    ใช้สำหรับไฟล์ที่มีชื่อโรงเรียน/วิทยาลัยอยู่ใน title row ก่อน header
    เช่น students-loeitech.xlsx มี "วิทยาลัยเทคนิคเลย" ใน Row 2
    """
    for row in all_rows[:header_idx]:
        for cell in row:
            if cell is None:
                continue
            text = str(cell).strip()
            if len(text) < 4:
                continue
            if any(text.startswith(p) for p in _SCHOOL_KNOWN_PREFIXES):
                return text
    return ""


def smart_parse_excel(content: bytes) -> dict:
    """
    อ่าน Excel แบบ smart:
    1. อ่านทุกแถว
    2. auto-detect header row
    3. map columns
    4. return metadata + preview rows
    """
    sheet_name, all_rows = _read_excel_raw(content)

    if not all_rows:
        return {
            "total_rows": 0, "header_row_index": 0,
            "headers": [], "column_mapping": {},
            "preview_raw": [], "preview_mapped": [],
            "sheet_name": sheet_name,
            "detected_school_name": "",
        }

    # 1. Detect header row
    header_idx = auto_detect_header_row(all_rows)
    headers = [str(c).strip() if c is not None else "" for c in all_rows[header_idx]]

    # 1.5 หาชื่อสถานศึกษาจาก title rows (ก่อน header)
    detected_school_name = _detect_school_from_title_rows(all_rows, header_idx)

    # 2. Map columns
    col_map = map_columns(headers)

    # 2.5 Data-driven fix: หา column ที่มีข้อมูล 13 หลักจริงๆ → assign เป็น national_id
    # (ใช้แก้กรณี SGS ที่มี 2 column ชื่อเหมือนกัน "เลขประจำตัวนักเรียน")
    data_rows = all_rows[header_idx + 1:]
    data_rows = [r for r in data_rows if any(c is not None and str(c).strip() for c in r)]

    sample_rows = data_rows[:10]
    thirteen_digit_cols: dict[int, int] = {}  # col_idx → count of valid national_id values

    for row in sample_rows:
        for ci, val in enumerate(row):
            if val is None:
                continue
            normalized, err = normalize_national_id(str(val))
            if normalized and not err:  # valid Thai ID หรือ G-Code
                thirteen_digit_cols[ci] = thirteen_digit_cols.get(ci, 0) + 1

    if thirteen_digit_cols:
        # เลือก col ที่มี 13-digit บ่อยที่สุด
        best_ni_col = max(thirteen_digit_cols, key=lambda x: thirteen_digit_cols[x])
        # ถ้า col นั้นยังไม่ถูก assign เป็น national_id → assign
        if col_map.get("national_id") != best_ni_col:
            displaced_field = None
            displaced_old_idx = None
            # ย้าย field อื่นที่อาจครอง col นี้ออกก่อน
            for field, idx in list(col_map.items()):
                if idx == best_ni_col and field != "national_id":
                    displaced_field = field
                    displaced_old_idx = idx
                    col_map[field] = None
                    break
            col_map["national_id"] = best_ni_col

            # ── ถ้า student_code ถูก displace → หา duplicate col ชื่อเดียวกันที่เหลืออยู่ ──
            # กรณี SGS: มี 2 col ชื่อ "เลขประจำตัวนักเรียน" col แรก=เลขปชช, col หลัง=running id
            if displaced_field == "student_code" and col_map["student_code"] is None:
                ni_header = headers[best_ni_col] if best_ni_col < len(headers) else ""
                assigned_cols = set(v for v in col_map.values() if v is not None)
                for ci, h in enumerate(headers):
                    if ci == best_ni_col or ci in assigned_cols:
                        continue
                    # หา col อื่นที่มีชื่อเหมือนกันหรือคล้ายกัน และมีตัวเลขสั้นๆ ในข้อมูล
                    if _similarity(h, ni_header) > 0.7 or _similarity(h, "เลขประจำตัวนักเรียน") > 0.7:
                        # ตรวจว่าข้อมูลใน col นี้ไม่ใช่ 13 digits
                        short_count = sum(
                            1 for row in sample_rows
                            if row[ci] is not None and 1 <= len(re.sub(r"\D", "", str(row[ci]))) <= 12
                        )
                        if short_count > len(sample_rows) // 2:
                            col_map["student_code"] = ci
                            break

    # 4. Preview 5 แถวแรก
    # preview_raw เก็บเป็น list[str] ตาม col index — ไม่ใช้ header name เป็น key
    # เพื่อรองรับกรณีที่มี column ชื่อซ้ำกัน (เช่น ไฟล์ SGS/เชียงกลม)
    preview_raw = []
    preview_mapped = []
    for row in data_rows[:5]:
        raw = [str(row[i]).strip() if i < len(row) and row[i] is not None else "" for i in range(len(headers))]
        mapped = _apply_mapping(row, headers, col_map)
        preview_raw.append(raw)
        preview_mapped.append(mapped)

    return {
        "total_rows": len(data_rows),
        "header_row_index": header_idx,
        "sheet_name": sheet_name,
        "headers": headers,
        "column_mapping": {field: ({"col_index": idx, "col_name": headers[idx]} if idx is not None else None)
                           for field, idx in col_map.items()},
        "preview_raw": preview_raw,
        "preview_mapped": preview_mapped,
        "detected_school_name": detected_school_name,
    }


def _apply_mapping(row: list[Any], headers: list[str], col_map: dict[str, int | None]) -> dict:
    """แปลงแถวข้อมูลตาม mapping → dict ที่มี field name ถูกต้อง"""
    def get(field: str) -> str:
        idx = col_map.get(field)
        if idx is None or idx >= len(row):
            return ""
        v = row[idx]
        return str(v).strip() if v is not None else ""

    prefix_raw = get("prefix")
    title = normalize_title(prefix_raw) if prefix_raw else ""
    gender_raw = get("gender")

    # Derive gender: ถ้ามี col gender ใช้เลย ไม่ก็ derive จาก title ที่ normalize แล้ว
    if gender_raw:
        gender = normalize_gender(gender_raw)
    elif title:
        gender = detect_gender_from_title(title)
    else:
        gender = "ไม่ระบุ"

    # Merge grade + classroom
    grade = get("grade")
    classroom_raw = get("classroom")

    # Case A: grade แบบ loeitech `ปวช.1/1` (รวมอยู่ใน 1 column, classroom col เป็น numeric id)
    # ในกรณีนี้ classroom_raw จะเป็นตัวเลขของ group id ไม่ใช่ ห้อง
    # Case B: เชียงกลม มีแยก grade=`อ.2` + classroom=`1`
    # ต้องรวมเป็น `อ.2/1`

    # ถ้า grade มี "/" อยู่แล้ว (เช่น "ปวช.1/1") → ใช้เป็น classroom เลย ไม่ต้อง append อีก
    if grade and "/" in grade:
        classroom = grade          # เช่น "ปวช.1/1"
        grade = grade.split("/")[0].strip()  # แค่ "ปวช.1"
    elif grade and classroom_raw:
        if classroom_raw.isdigit() or (len(classroom_raw) <= 3 and classroom_raw.strip()):
            # เชียงกลม: ชั้น=อ.2 + ห้อง=1 → "อ.2/1"
            classroom = f"{grade}/{classroom_raw}"
        else:
            classroom = classroom_raw  # ใช้ classroom_raw ตรงๆ
    elif classroom_raw:
        classroom = classroom_raw
    else:
        classroom = grade

    student_code_val = get("student_code")
    national_id_val = get("national_id")

    # ── ไม่ swap อัตโนมัติ: ให้ admin ตรวจสอบ mapping ใน UI เอง ──
    # การ swap อัตโนมัติมีความเสี่ยงสูงที่จะทำให้ student_code และ national_id
    # สลับกันในฐานข้อมูล ซึ่งทำให้ข้อมูลเสียหายถาวร

    return {
        "student_code": student_code_val,
        "national_id": national_id_val,
        "title": title,
        "first_name": get("first_name"),
        "last_name": get("last_name"),
        "gender": gender,
        "birthdate": get("birthdate"),
        "grade": grade,
        "classroom": classroom,
        "school_name": get("school_name"),
        "school_code": get("school_code"),
        "status": get("status"),
    }


def parse_excel(content: bytes) -> list[dict]:
    """Legacy: parse Excel แบบเดิม (ต้องใช้ template header ที่ถูกต้อง)"""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    result = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        result.append({headers[i]: (str(row[i]).strip() if row[i] is not None else "") for i in range(len(headers))})
    return result


# ─── Smart Bulk Import ────────────────────────────────────────────────────────

async def smart_bulk_import_students(
    db: AsyncSession,
    content: bytes,
    school_id: int,
    col_map_override: dict[str, int | None] | None = None,
) -> dict:
    """
    Smart import: รับ Excel binary + school_id
    - auto-detect header
    - apply column mapping (หรือใช้ override จาก frontend ถ้ามี)
    - insert/update Student records
    - encrypt national_id ด้วย AES-256
    """
    sheet_name, all_rows = _read_excel_raw(content)

    if not all_rows:
        return {"created": 0, "updated": 0, "errors": [], "skipped": 0}

    # Detect header
    header_idx = auto_detect_header_row(all_rows)
    headers = [str(c).strip() if c is not None else "" for c in all_rows[header_idx]]

    # Use provided mapping or auto-detect
    if col_map_override:
        # col_map_override format: {field: col_index}
        col_map = col_map_override
    else:
        col_map = map_columns(headers)

    # Data rows
    data_rows = all_rows[header_idx + 1:]
    data_rows = [r for r in data_rows if any(c is not None and str(c).strip() for c in r)]

    created, updated, skipped, errors = 0, 0, 0, []
    nid_warnings: list[dict] = []  # แถวที่มี national_id แต่ไม่ครบ 13 หลัก

    for row_offset, row in enumerate(data_rows):
        row_num = header_idx + 2 + row_offset  # actual Excel row number
        try:
            mapped = _apply_mapping(row, headers, col_map)

            student_code = mapped["student_code"].strip()
            first_name = mapped["first_name"].strip()
            last_name = mapped["last_name"].strip()

            # Validation
            if not student_code:
                errors.append({"row": row_num, "reason": "ไม่พบรหัสนักเรียน (student_code ว่าง)"})
                continue
            if not first_name:
                errors.append({"row": row_num, "reason": f"student_code={student_code}: ชื่อว่าง"})
                continue

            # Check status — ข้ามนักเรียนที่ออกแล้ว
            status = mapped.get("status", "")
            if status and status not in ("กำลังศึกษา", ""):
                skipped += 1
                continue

            # ข้ามระดับชั้นอนุบาล (อ.1 อ.2 อ.3) — ไม่อยู่ในขอบเขตการประเมิน
            grade_val = mapped.get("grade", "").strip()
            if grade_val.startswith("อ.") or grade_val in ("อนุบาล 1", "อนุบาล 2", "อนุบาล 3", "อนุบาล1", "อนุบาล2", "อนุบาล3"):
                skipped += 1
                continue

            # Parse birthdate
            birthdate = parse_thai_date(mapped["birthdate"])

            # Find existing student
            result = await db.execute(select(Student).where(Student.student_code == student_code))
            existing = result.scalar_one_or_none()

            # ตรวจ student_code ว่าดูเหมือน national_id หรือไม่ (mapping สลับกัน)
            sc_digits = re.sub(r"\D", "", student_code)
            if re.match(r"^\d{13}$", sc_digits) or re.match(r"^[Gg]\d{12}$", student_code):
                errors.append({
                    "row": row_num,
                    "reason": (
                        f"student_code='{student_code}' ดูเหมือนเลขบัตรประชาชน (13 หลัก) "
                        f"— กรุณาตรวจสอบ Column Mapping ว่า 'รหัสนักเรียน' และ 'เลขบัตรปชช.' ไม่สลับกัน"
                    ),
                })
                continue  # ไม่ import แถวนี้

            national_id_raw = mapped["national_id"].strip()
            national_id, nid_err = normalize_national_id(national_id_raw)
            if nid_err:
                nid_warnings.append({
                    "row": row_num,
                    "student_code": student_code,
                    "name": f"{first_name} {last_name}".strip(),
                    "reason": f"{nid_err} — ข้ามการบันทึกเลขบัตรปชช.",
                })

            if existing:
                # Upsert: update fields ที่เปลี่ยน
                if first_name:
                    existing.first_name = first_name
                if last_name:
                    existing.last_name = last_name
                if mapped.get("title"):
                    existing.title = mapped["title"]
                if mapped["gender"]:
                    existing.gender = mapped["gender"]
                if mapped["grade"]:
                    existing.grade = mapped["grade"]
                if mapped["classroom"]:
                    existing.classroom = mapped["classroom"]
                if school_id:
                    existing.school_id = school_id
                if birthdate:
                    existing.birthdate = birthdate
                if national_id:
                    existing.national_id = encrypt_pii(national_id)
                    existing.national_id_hash = hash_pii(national_id)
                updated += 1
            else:
                stu = Student(
                    student_code=student_code,
                    title=mapped.get("title") or None,
                    first_name=first_name,
                    last_name=last_name or "",
                    gender=mapped["gender"] or None,
                    grade=mapped["grade"] or None,
                    classroom=mapped["classroom"] or None,
                    school_id=school_id,
                    birthdate=birthdate,
                    is_active=True,
                )
                if national_id:
                    stu.national_id = encrypt_pii(national_id)
                    stu.national_id_hash = hash_pii(national_id)
                db.add(stu)
                await db.flush()  # get stu.id before creating user

                # Auto-create User record for student login
                user = User(student_id=stu.id, role="student", school_id=school_id)
                db.add(user)
                created += 1

        except Exception as e:
            errors.append({"row": row_num, "reason": str(e)})

    await db.commit()
    return {
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "errors": errors,
        "nid_warnings": nid_warnings,
        "total_processed": len(data_rows),
    }


# ─── Legacy functions (kept for compatibility) ────────────────────────────────

async def bulk_import_students(db: AsyncSession, rows: list[dict]) -> dict:
    """Legacy: ใช้กับ CSV template เดิม"""
    created, updated, errors = 0, 0, []

    for i, row in enumerate(rows, start=2):
        try:
            student_code = row.get("student_code", "").strip()
            if not student_code:
                errors.append({"row": i, "reason": "student_code ว่าง"})
                continue

            school_id = int(row.get("school_id", 0)) if row.get("school_id") else None

            result = await db.execute(select(Student).where(Student.student_code == student_code))
            existing = result.scalar_one_or_none()

            if existing:
                if row.get("first_name"): existing.first_name = row["first_name"]
                if row.get("last_name"):  existing.last_name = row["last_name"]
                if row.get("gender"):     existing.gender = row["gender"]
                if row.get("grade"):      existing.grade = row["grade"]
                if row.get("classroom"):  existing.classroom = row["classroom"]
                if school_id:             existing.school_id = school_id
                if row.get("birthdate"):
                    try:
                        y, m, d = row["birthdate"].split("-")
                        existing.birthdate = date(int(y), int(m), int(d))
                    except Exception:
                        pass
                updated += 1
            else:
                national_id = row.get("national_id", "").strip()
                stu = Student(
                    student_code=student_code,
                    first_name=row.get("first_name", ""),
                    last_name=row.get("last_name", ""),
                    gender=row.get("gender") or None,
                    grade=row.get("grade") or None,
                    classroom=row.get("classroom") or None,
                    school_id=school_id,
                    is_active=True,
                )
                if national_id:
                    stu.national_id = encrypt_pii(national_id)
                    stu.national_id_hash = hash_pii(national_id)
                if row.get("birthdate"):
                    try:
                        y, m, d2 = row["birthdate"].split("-")
                        stu.birthdate = date(int(y), int(m), int(d2))
                    except Exception:
                        pass
                db.add(stu)
                created += 1

        except Exception as e:
            errors.append({"row": i, "reason": str(e)})

    await db.commit()
    return {"created": created, "updated": updated, "errors": errors}


async def bulk_import_schools(db: AsyncSession, rows: list[dict]) -> dict:
    """Legacy: ใช้กับ CSV template เดิม"""
    created, updated, errors = 0, 0, []

    for i, row in enumerate(rows, start=2):
        try:
            school_type = (row.get("school_type") or "").strip() or None
            raw_name = row.get("name", "").strip()
            if not raw_name:
                errors.append({"row": i, "reason": "name ว่าง"})
                continue
            name = normalize_school_name(raw_name, school_type)

            district_id = int(row.get("district_id", 0)) if row.get("district_id") else None
            if not district_id:
                errors.append({"row": i, "reason": "district_id ว่าง"})
                continue

            result = await db.execute(select(School).where(School.name == name))
            existing = result.scalar_one_or_none()

            if existing:
                existing.district_id = district_id
                if school_type: existing.school_type = school_type
                updated += 1
            else:
                db.add(School(
                    name=name,
                    district_id=district_id,
                    school_type=school_type,
                ))
                created += 1

        except Exception as e:
            errors.append({"row": i, "reason": str(e)})

    await db.commit()
    return {"created": created, "updated": updated, "errors": errors}
